#!/usr/bin/env python3
"""Generate SQL to seed 2026 SDCC events from the coordination spreadsheet."""
import openpyxl
import re
import sys

XLSX = '/Users/tony/Downloads/Comic-Con Purchasing Train Coordination Sheet.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

def badge(raw):
    if raw and str(raw).upper().startswith('JUNIOR'):
        return 'JUNIOR'
    return 'ADULT'

def b(val):
    return 1 if val else 0

def pur(val):
    try:
        return 1 if float(val or 0) > 0 else 0
    except (TypeError, ValueError):
        return 0

def esc(s):
    return str(s or '').replace("'", "''").strip()

def parse_return_reg():
    ws = wb['2026 - Return Reg']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3: continue
        first = row[0]
        if not first or str(first).strip() in ('Unconfirmed', ''):
            continue
        first = str(first).strip()
        last  = str(row[1] or '').strip()
        mid   = str(row[2] or '').strip()
        bt    = badge(row[4])
        ret_elig   = b(row[3])
        req_pv     = b(row[5])
        req_thu    = b(row[6])
        req_fri    = b(row[7])
        req_sat    = b(row[8])
        req_sun    = b(row[9])
        coordinator = esc(row[12])
        pur_pv     = pur(row[13])
        pur_thu    = pur(row[14])
        pur_fri    = pur(row[15])
        pur_sat    = pur(row[16])
        pur_sun    = pur(row[17])
        who_pur    = esc(row[18])
        paid       = b(row[21])
        notes_raw  = str(row[22] or '').strip()

        # Parse sponsor out of notes
        sponsor = ''
        notes = notes_raw
        m = re.search(r'[Ss]ponsor(?:ed\s+by)?:?\s*([^,;]+)', notes_raw)
        if m:
            sponsor = m.group(1).strip()
            notes = re.sub(r'[Ss]ponsor(?:ed\s+by)?:?\s*[^,;]+[,;]?\s*', '', notes_raw).strip()

        has_data = any([req_pv, req_thu, req_fri, req_sat, req_sun,
                        pur_pv, pur_thu, pur_fri, pur_sat, pur_sun, coordinator])
        if not has_data:
            continue

        rows.append(dict(
            first_name=first, last_name=last, member_id=mid,
            badge_type=bt, return_eligible=ret_elig, sponsor=esc(sponsor), notes=esc(notes),
            req_preview=req_pv, req_thu=req_thu, req_fri=req_fri, req_sat=req_sat, req_sun=req_sun,
            purchasing_coordinator=coordinator,
            pur_preview=pur_pv, pur_thu=pur_thu, pur_fri=pur_fri, pur_sat=pur_sat, pur_sun=pur_sun,
            who_purchased=who_pur, paid=paid,
            sort_order=len(rows) + 1,
        ))
    return rows

def parse_open_reg():
    ws = wb['2026 - Open Reg']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3: continue
        first = row[0]
        if not first or str(first).strip() in ('Unconfirmed', ''):
            continue
        first = str(first).strip()
        last  = str(row[1] or '').strip()
        mid   = str(row[2] or '').strip()
        bt    = badge(row[3])
        # Open Reg has no return_eligible column
        req_pv     = b(row[4])
        req_thu    = b(row[5])
        req_fri    = b(row[6])
        req_sat    = b(row[7])
        req_sun    = b(row[8])
        coordinator = esc(row[11])
        pur_pv     = pur(row[12])
        pur_thu    = pur(row[13])
        pur_fri    = pur(row[14])
        pur_sat    = pur(row[15])
        pur_sun    = pur(row[16])
        who_pur    = esc(row[17])
        paid       = b(row[20])
        notes      = esc(row[21])

        has_data = any([req_pv, req_thu, req_fri, req_sat, req_sun,
                        pur_pv, pur_thu, pur_fri, pur_sat, pur_sun, coordinator])
        if not has_data:
            continue

        rows.append(dict(
            first_name=first, last_name=last, member_id=mid,
            badge_type=bt, return_eligible=0, sponsor='', notes=notes,
            req_preview=req_pv, req_thu=req_thu, req_fri=req_fri, req_sat=req_sat, req_sun=req_sun,
            purchasing_coordinator=coordinator,
            pur_preview=pur_pv, pur_thu=pur_thu, pur_fri=pur_fri, pur_sat=pur_sat, pur_sun=pur_sun,
            who_purchased=who_pur, paid=paid,
            sort_order=len(rows) + 1,
        ))
    return rows

def participant_sql(event_id_placeholder, p):
    return (
        f"INSERT INTO participants "
        f"(event_id, first_name, last_name, member_id, badge_type, return_eligible, sponsor, notes, "
        f"req_preview, req_thu, req_fri, req_sat, req_sun, sort_order, "
        f"purchasing_coordinator, purchasing_claimed_by, "
        f"pur_preview, pur_thu, pur_fri, pur_sat, pur_sun, who_purchased, paid) VALUES "
        f"({event_id_placeholder}, '{esc(p['first_name'])}', '{esc(p['last_name'])}', "
        f"'{p['member_id']}', '{p['badge_type']}', {p['return_eligible']}, "
        f"'{p['sponsor']}', '{p['notes']}', "
        f"{p['req_preview']}, {p['req_thu']}, {p['req_fri']}, {p['req_sat']}, {p['req_sun']}, "
        f"{p['sort_order']}, "
        f"'{p['purchasing_coordinator']}', '{p['purchasing_coordinator']}', "
        f"{p['pur_preview']}, {p['pur_thu']}, {p['pur_fri']}, {p['pur_sat']}, {p['pur_sun']}, "
        f"'{p['who_purchased']}', {p['paid']});"
    )

return_rows = parse_return_reg()
open_rows   = parse_open_reg()

print(f"Return Reg: {len(return_rows)} participants", file=sys.stderr)
print(f"Open Reg:   {len(open_rows)} participants", file=sys.stderr)

lines = []
lines.append("-- 2026 Return Reg event")
lines.append(
    "INSERT INTO events (year, name, reg_type, status, "
    "price_preview_adult, price_thu_adult, price_fri_adult, price_sat_adult, price_sun_adult, "
    "price_preview_junior, price_thu_junior, price_fri_junior, price_sat_junior, price_sun_junior, "
    "access_token) VALUES "
    "(2026, 'SDCC 2026 Return Reg', 'return', 'complete', "
    "6400, 8500, 8500, 8500, 6400, 3800, 4400, 4400, 4400, 3800, '');"
)
lines.append("")
lines.append("-- 2026 Open Reg event")
lines.append(
    "INSERT INTO events (year, name, reg_type, status, "
    "price_preview_adult, price_thu_adult, price_fri_adult, price_sat_adult, price_sun_adult, "
    "price_preview_junior, price_thu_junior, price_fri_junior, price_sat_junior, price_sun_junior, "
    "access_token) VALUES "
    "(2026, 'SDCC 2026 Open Reg', 'open', 'complete', "
    "6400, 8500, 8500, 8500, 6400, 3800, 4400, 4400, 4400, 3800, '');"
)
lines.append("")

lines.append("-- Return Reg participants (event inserted last, so id = last_insert_rowid() won't work)")
lines.append("-- We use a subquery to get the event id by name")
for p in return_rows:
    lines.append(participant_sql(
        "(SELECT id FROM events WHERE name = 'SDCC 2026 Return Reg')", p
    ))

lines.append("")
lines.append("-- Open Reg participants")
for p in open_rows:
    lines.append(participant_sql(
        "(SELECT id FROM events WHERE name = 'SDCC 2026 Open Reg')", p
    ))

print('\n'.join(lines))
