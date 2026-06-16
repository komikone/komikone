#!/usr/bin/env python3
"""Parse the SDCC spreadsheet and generate SQL to seed historical years into D1."""

import openpyxl
import secrets
import sys

XLSX = '/Users/tony/Downloads/Comic-Con Purchasing Train Coordination Sheet.xlsx'

wb = openpyxl.load_workbook(XLSX, data_only=True)


def normalize_badge(v):
    if not v:
        return 'ADULT'
    s = str(v).strip().upper()
    if any(k in s for k in ('JUNIOR', 'YOUTH', 'MILITARY', 'SENIOR')):
        return 'JUNIOR'
    return 'ADULT'


def to_bool(v):
    if v is None or v is False:
        return False
    if v is True:
        return True
    if isinstance(v, (int, float)):
        return v > 0
    return False


def to_cents(v):
    """Dollar float → cents int. '--' or None → 0."""
    if v is None or v == '--' or v == '':
        return 0
    try:
        return int(round(float(v) * 100))
    except (TypeError, ValueError):
        return 0


def sql_str(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_bool(v):
    return '1' if v else '0'


# ── Column parsers per sheet format ──────────────────────────────────────────

def parse_2020(row):
    """Row format from the '2020' sheet (Returning Attendees)."""
    # 0:first 1:last 2:member_id 3:badge 4:req_pre 5:req_thu 6:req_fri
    # 7:req_sat 8:req_sun 9:who_purchased 10:pur_pre 11:pur_thu 12:pur_fri
    # 13:pur_sat 14:pur_sun 15:gaps 16:total 17:paid
    if not row[0]:
        return None
    return {
        'first_name': str(row[0]).strip(),
        'last_name': str(row[1] or '').strip(),
        'member_id': str(row[2] or '').strip().upper(),
        'badge_type': normalize_badge(row[3]),
        'req_preview': to_bool(row[4]), 'req_thu': to_bool(row[5]),
        'req_fri': to_bool(row[6]), 'req_sat': to_bool(row[7]), 'req_sun': to_bool(row[8]),
        'who_purchased': str(row[9] or '').strip(),
        'pur_preview': to_bool(row[10]), 'pur_thu': to_bool(row[11]),
        'pur_fri': to_bool(row[12]), 'pur_sat': to_bool(row[13]), 'pur_sun': to_bool(row[14]),
        'purchase_total': to_cents(row[16]),
        'paid': to_bool(row[17]),
        'notes': '',
        'return_eligible': True,
    }


def parse_2023(row):
    """Row format from 2023/2024 sheets (no 'Purchasing' column)."""
    # 0:first 1:last 2:member_id 3:badge 4:req_pre 5:req_thu 6:req_fri
    # 7:req_sat 8:req_sun 9:empty 10:pur_pre 11:pur_thu 12:pur_fri
    # 13:pur_sat 14:pur_sun 15:who_purchased 16:gaps 17:total 18:paid 19:notes
    if not row[0]:
        return None
    return {
        'first_name': str(row[0]).strip(),
        'last_name': str(row[1] or '').strip(),
        'member_id': str(row[2] or '').strip().upper(),
        'badge_type': normalize_badge(row[3]),
        'req_preview': to_bool(row[4]), 'req_thu': to_bool(row[5]),
        'req_fri': to_bool(row[6]), 'req_sat': to_bool(row[7]), 'req_sun': to_bool(row[8]),
        'pur_preview': to_bool(row[10]), 'pur_thu': to_bool(row[11]),
        'pur_fri': to_bool(row[12]), 'pur_sat': to_bool(row[13]), 'pur_sun': to_bool(row[14]),
        'who_purchased': str(row[15] or '').strip(),
        'purchase_total': to_cents(row[17]),
        'paid': to_bool(row[18]),
        'notes': str(row[19] or '').strip(),
        'return_eligible': False,
    }


def parse_2025(row):
    """Row format from 2025 sheets (adds 'Purchasing' column before purchased days)."""
    # 0:first 1:last 2:member_id 3:badge 4:req_pre 5:req_thu 6:req_fri
    # 7:req_sat 8:req_sun 9:empty 10:purchasing_bool 11:empty
    # 12:pur_pre 13:pur_thu 14:pur_fri 15:pur_sat 16:pur_sun
    # 17:who_purchased 18:gaps 19:total 20:paid 21:notes
    if not row[0]:
        return None
    return {
        'first_name': str(row[0]).strip(),
        'last_name': str(row[1] or '').strip(),
        'member_id': str(row[2] or '').strip().upper(),
        'badge_type': normalize_badge(row[3]),
        'req_preview': to_bool(row[4]), 'req_thu': to_bool(row[5]),
        'req_fri': to_bool(row[6]), 'req_sat': to_bool(row[7]), 'req_sun': to_bool(row[8]),
        'pur_preview': to_bool(row[12]), 'pur_thu': to_bool(row[13]),
        'pur_fri': to_bool(row[14]), 'pur_sat': to_bool(row[15]), 'pur_sun': to_bool(row[16]),
        'who_purchased': str(row[17] or '').strip(),
        'purchase_total': to_cents(row[19]),
        'paid': to_bool(row[20]),
        'notes': str(row[21] or '').strip(),
        'return_eligible': False,
    }


# ── Sheet definitions ─────────────────────────────────────────────────────────

SHEETS = [
    {
        'name': '2020', 'year': 2020, 'reg_type': 'return',
        'event_name': 'SDCC 2020 Return Reg',
        'data_start': 4, 'parser': parse_2020,
        'prices': {
            'price_preview_adult': 0, 'price_thu_adult': 6850, 'price_fri_adult': 6850,
            'price_sat_adult': 6850, 'price_sun_adult': 4600,
            'price_preview_junior': 0, 'price_thu_junior': 3400, 'price_fri_junior': 3400,
            'price_sat_junior': 3400, 'price_sun_junior': 2300,
        },
    },
    {
        'name': '2023 - Return Reg', 'year': 2023, 'reg_type': 'return',
        'event_name': 'SDCC 2023 Return Reg',
        'data_start': 4, 'parser': parse_2023,
        'ret_eligible': True,
        'prices': {
            'price_preview_adult': 5500, 'price_thu_adult': 7500, 'price_fri_adult': 7500,
            'price_sat_adult': 7500, 'price_sun_adult': 5000,
            'price_preview_junior': 0, 'price_thu_junior': 3700, 'price_fri_junior': 3700,
            'price_sat_junior': 3700, 'price_sun_junior': 2500,
        },
    },
    {
        'name': '2023 - Open Reg', 'year': 2023, 'reg_type': 'open',
        'event_name': 'SDCC 2023 Open Reg',
        'data_start': 4, 'parser': parse_2023,
        'prices': {
            'price_preview_adult': 5500, 'price_thu_adult': 7500, 'price_fri_adult': 7500,
            'price_sat_adult': 7500, 'price_sun_adult': 5000,
            'price_preview_junior': 0, 'price_thu_junior': 3700, 'price_fri_junior': 3700,
            'price_sat_junior': 3700, 'price_sun_junior': 2500,
        },
    },
    {
        'name': '2024 - Return Reg', 'year': 2024, 'reg_type': 'return',
        'event_name': 'SDCC 2024 Return Reg',
        'data_start': 4, 'parser': parse_2023,
        'ret_eligible': True,
        'prices': {
            'price_preview_adult': 5900, 'price_thu_adult': 7900, 'price_fri_adult': 7900,
            'price_sat_adult': 7900, 'price_sun_adult': 5000,
            'price_preview_junior': 0, 'price_thu_junior': 3700, 'price_fri_junior': 3700,
            'price_sat_junior': 3700, 'price_sun_junior': 2500,
        },
    },
    {
        'name': '2024 - Open Reg', 'year': 2024, 'reg_type': 'open',
        'event_name': 'SDCC 2024 Open Reg',
        'data_start': 4, 'parser': parse_2023,
        'prices': {
            'price_preview_adult': 5900, 'price_thu_adult': 7900, 'price_fri_adult': 7900,
            'price_sat_adult': 7900, 'price_sun_adult': 5000,
            'price_preview_junior': 0, 'price_thu_junior': 3700, 'price_fri_junior': 3700,
            'price_sat_junior': 3700, 'price_sun_junior': 2500,
        },
    },
    {
        'name': '2025 - Return Reg', 'year': 2025, 'reg_type': 'return',
        'event_name': 'SDCC 2025 Return Reg',
        'data_start': 4, 'parser': parse_2025,
        'ret_eligible': True,
        'prices': {
            'price_preview_adult': 6100, 'price_thu_adult': 8000, 'price_fri_adult': 8000,
            'price_sat_adult': 8000, 'price_sun_adult': 6000,
            'price_preview_junior': 6100, 'price_thu_junior': 4000, 'price_fri_junior': 4000,
            'price_sat_junior': 4000, 'price_sun_junior': 3000,
        },
    },
    {
        'name': '2025 - Open Reg', 'year': 2025, 'reg_type': 'open',
        'event_name': 'SDCC 2025 Open Reg',
        'data_start': 4, 'parser': parse_2025,
        'prices': {
            'price_preview_adult': 6100, 'price_thu_adult': 8000, 'price_fri_adult': 8000,
            'price_sat_adult': 8000, 'price_sun_adult': 6000,
            'price_preview_junior': 6100, 'price_thu_junior': 4000, 'price_fri_junior': 4000,
            'price_sat_junior': 4000, 'price_sun_junior': 3000,
        },
    },
]

# ── Generate SQL ──────────────────────────────────────────────────────────────

lines = [
    '-- Historical SDCC data seed',
    '-- Generated by seed_historical.py',
    '',
]

for sheet_cfg in SHEETS:
    ws = wb[sheet_cfg['name']]
    year = sheet_cfg['year']
    reg_type = sheet_cfg['reg_type']
    event_name = sheet_cfg['event_name']
    p = sheet_cfg['prices']
    ret_eligible_override = sheet_cfg.get('ret_eligible', False)
    parser = sheet_cfg['parser']

    token = secrets.token_urlsafe(24)

    # Insert event
    lines.append(f'-- ── {event_name} ──')
    lines.append(
        f"INSERT INTO events (year, name, reg_type, status, access_token, "
        f"price_preview_adult, price_thu_adult, price_fri_adult, price_sat_adult, price_sun_adult, "
        f"price_preview_junior, price_thu_junior, price_fri_junior, price_sat_junior, price_sun_junior) "
        f"VALUES ({year}, {sql_str(event_name)}, {sql_str(reg_type)}, 'complete', {sql_str(token)}, "
        f"{p['price_preview_adult']}, {p['price_thu_adult']}, {p['price_fri_adult']}, {p['price_sat_adult']}, {p['price_sun_adult']}, "
        f"{p['price_preview_junior']}, {p['price_thu_junior']}, {p['price_fri_junior']}, {p['price_sat_junior']}, {p['price_sun_junior']}"
        f");"
    )
    lines.append('')

    # Parse participants
    participants = []
    for row in ws.iter_rows(min_row=sheet_cfg['data_start'], values_only=True):
        p_data = parser(row)
        if p_data is None:
            break
        if ret_eligible_override:
            p_data['return_eligible'] = True
        participants.append(p_data)

    if not participants:
        lines.append(f'-- (no participants found in {sheet_cfg["name"]})')
        lines.append('')
        continue

    for sort_order, p_data in enumerate(participants):
        # Derive all_purchased: every requested day was purchased
        req_days = ['preview', 'thu', 'fri', 'sat', 'sun']
        any_req = any(p_data[f'req_{d}'] for d in req_days)
        all_pur = any_req and all(
            (not p_data[f'req_{d}']) or p_data[f'pur_{d}']
            for d in req_days
        )

        lines.append(
            f"INSERT INTO participants "
            f"(event_id, first_name, last_name, member_id, badge_type, return_eligible, "
            f"req_preview, req_thu, req_fri, req_sat, req_sun, "
            f"pur_preview, pur_thu, pur_fri, pur_sat, pur_sun, "
            f"who_purchased, paid, notes, sort_order) "
            f"VALUES ("
            f"(SELECT id FROM events WHERE name = {sql_str(event_name)} LIMIT 1), "
            f"{sql_str(p_data['first_name'])}, {sql_str(p_data['last_name'])}, "
            f"{sql_str(p_data['member_id'])}, {sql_str(p_data['badge_type'])}, "
            f"{sql_bool(p_data['return_eligible'])}, "
            f"{sql_bool(p_data['req_preview'])}, {sql_bool(p_data['req_thu'])}, "
            f"{sql_bool(p_data['req_fri'])}, {sql_bool(p_data['req_sat'])}, {sql_bool(p_data['req_sun'])}, "
            f"{sql_bool(p_data['pur_preview'])}, {sql_bool(p_data['pur_thu'])}, "
            f"{sql_bool(p_data['pur_fri'])}, {sql_bool(p_data['pur_sat'])}, {sql_bool(p_data['pur_sun'])}, "
            f"{sql_str(p_data['who_purchased'])}, "
            f"{sql_bool(p_data['paid'])}, {sql_str(p_data['notes'])}, {sort_order}"
            f");"
        )
    lines.append('')

sql = '\n'.join(lines)
print(sql)

# Also print summary
total_p = 0
print('', file=sys.stderr)
print('Summary:', file=sys.stderr)
for sheet_cfg in SHEETS:
    ws = wb[sheet_cfg['name']]
    count = 0
    for row in ws.iter_rows(min_row=sheet_cfg['data_start'], values_only=True):
        if sheet_cfg['parser'](row) is not None:
            count += 1
        else:
            break
    total_p += count
    print(f"  {sheet_cfg['name']}: {count} participants", file=sys.stderr)
print(f'  Total: {total_p} participants across {len(SHEETS)} events', file=sys.stderr)
